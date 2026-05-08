import os, base64, secrets, hashlib, requests, smtplib, json, io
from datetime import datetime, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from functools import wraps
from flask import Flask, request, jsonify, render_template, session, redirect, send_file, Response
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import RealDictCursor
from contextlib import contextmanager

load_dotenv()
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(32))

# ══════════════════════════════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════════════════════════════
_raw_db = os.getenv("DATABASE_URL", "")
if _raw_db.startswith("postgresql://"):
    _raw_db = _raw_db.replace("postgresql://", "postgres://", 1)
DATABASE_URL = _raw_db

@contextmanager
def get_db():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL is not set.")
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor, sslmode="require")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def init_db():
    with get_db() as conn:
        cur = conn.cursor()
        # Users
        cur.execute("""CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY, email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL, name TEXT NOT NULL,
            business_name TEXT NOT NULL, phone TEXT DEFAULT '',
            mpesa_phone TEXT DEFAULT '', till_number TEXT DEFAULT '',
            paybill_number TEXT DEFAULT '', mpesa_shortcode TEXT DEFAULT '',
            mpesa_passkey TEXT DEFAULT '', role TEXT DEFAULT 'owner',
            owner_id INTEGER REFERENCES users(id),
            branch_id INTEGER, is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        # OTPs
        cur.execute("""CREATE TABLE IF NOT EXISTS otps (
            id SERIAL PRIMARY KEY, identifier TEXT NOT NULL,
            code TEXT NOT NULL, purpose TEXT NOT NULL,
            expires_at TIMESTAMP NOT NULL, used BOOLEAN DEFAULT FALSE
        )""")
        # Branches
        cur.execute("""CREATE TABLE IF NOT EXISTS branches (
            id SERIAL PRIMARY KEY, owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            name TEXT NOT NULL, location TEXT DEFAULT '',
            inventory_mode TEXT DEFAULT 'shared',
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        # Products
        cur.execute("""CREATE TABLE IF NOT EXISTS products (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            branch_id INTEGER REFERENCES branches(id) ON DELETE CASCADE,
            name TEXT NOT NULL, category TEXT DEFAULT 'General',
            barcode TEXT DEFAULT '', price INTEGER DEFAULT 0,
            wholesale_price INTEGER DEFAULT 0, bulk_price INTEGER DEFAULT 0,
            discount_pct INTEGER DEFAULT 0, discount_fixed INTEGER DEFAULT 0,
            vat_exempt BOOLEAN DEFAULT FALSE,
            stock INTEGER DEFAULT 0, reorder INTEGER DEFAULT 5,
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        # Customers
        cur.execute("""CREATE TABLE IF NOT EXISTS customers (
            id SERIAL PRIMARY KEY, owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            name TEXT NOT NULL, phone TEXT DEFAULT '', email TEXT DEFAULT '',
            customer_type TEXT DEFAULT 'retail', loyalty_points INTEGER DEFAULT 0,
            total_spent INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT NOW()
        )""")
        # Suppliers
        cur.execute("""CREATE TABLE IF NOT EXISTS suppliers (
            id SERIAL PRIMARY KEY, owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            name TEXT NOT NULL, phone TEXT DEFAULT '', email TEXT DEFAULT '',
            address TEXT DEFAULT '', balance_owed INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        # Purchase orders
        cur.execute("""CREATE TABLE IF NOT EXISTS purchase_orders (
            id SERIAL PRIMARY KEY, owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            branch_id INTEGER REFERENCES branches(id) ON DELETE SET NULL,
            supplier_id INTEGER REFERENCES suppliers(id) ON DELETE SET NULL,
            total_cost INTEGER DEFAULT 0, status TEXT DEFAULT 'received',
            note TEXT DEFAULT '', created_at TIMESTAMP DEFAULT NOW()
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS purchase_order_items (
            id SERIAL PRIMARY KEY,
            order_id INTEGER REFERENCES purchase_orders(id) ON DELETE CASCADE,
            product_id INTEGER REFERENCES products(id) ON DELETE SET NULL,
            name TEXT NOT NULL, qty INTEGER NOT NULL,
            unit_cost INTEGER NOT NULL, subtotal INTEGER NOT NULL
        )""")
        # Sales
        cur.execute("""CREATE TABLE IF NOT EXISTS sales (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            cashier_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
            customer_id INTEGER REFERENCES customers(id) ON DELETE SET NULL,
            branch_id INTEGER REFERENCES branches(id) ON DELETE SET NULL,
            txn_id TEXT NOT NULL, customer TEXT DEFAULT 'Walk-in',
            mpesa_code TEXT DEFAULT '', total_amount INTEGER DEFAULT 0,
            vat_amount INTEGER DEFAULT 0, discount INTEGER DEFAULT 0,
            price_tier TEXT DEFAULT 'retail', phone TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS sale_items (
            id SERIAL PRIMARY KEY, sale_id INTEGER REFERENCES sales(id) ON DELETE CASCADE,
            product_id INTEGER REFERENCES products(id) ON DELETE SET NULL,
            name TEXT NOT NULL, category TEXT DEFAULT '',
            price INTEGER NOT NULL, qty INTEGER NOT NULL, subtotal INTEGER NOT NULL
        )""")
        # Refunds
        cur.execute("""CREATE TABLE IF NOT EXISTS refunds (
            id SERIAL PRIMARY KEY, owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            sale_id INTEGER REFERENCES sales(id) ON DELETE SET NULL,
            txn_id TEXT, reason TEXT, note TEXT DEFAULT '',
            total_amount INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT NOW()
        )""")
        # Expenses
        cur.execute("""CREATE TABLE IF NOT EXISTS expenses (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            branch_id INTEGER REFERENCES branches(id) ON DELETE SET NULL,
            description TEXT NOT NULL, amount INTEGER NOT NULL,
            category TEXT DEFAULT 'General', note TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        # Subscriptions
        cur.execute("""CREATE TABLE IF NOT EXISTS subscriptions (
            id SERIAL PRIMARY KEY, owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE UNIQUE,
            status TEXT DEFAULT 'trial',
            plan TEXT DEFAULT 'basic',
            amount INTEGER DEFAULT 999,
            trial_ends_at TIMESTAMP,
            next_billing_date DATE,
            last_paid_at TIMESTAMP,
            checkout_id TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        # Reports
        cur.execute("""CREATE TABLE IF NOT EXISTS report_history (
            id SERIAL PRIMARY KEY, owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            report_type TEXT NOT NULL, delivered_via TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )""")
        # EOD
        cur.execute("""CREATE TABLE IF NOT EXISTS eod_reports (
            id SERIAL PRIMARY KEY, owner_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            branch_id INTEGER REFERENCES branches(id) ON DELETE SET NULL,
            report_date DATE NOT NULL, total_sales INTEGER DEFAULT 0,
            total_txns INTEGER DEFAULT 0, total_expenses INTEGER DEFAULT 0,
            vat_collected INTEGER DEFAULT 0, discounts_given INTEGER DEFAULT 0,
            net_profit INTEGER DEFAULT 0, created_at TIMESTAMP DEFAULT NOW()
        )""")

try:
    init_db()
    print("[DB] Tables ready")
except Exception as e:
    print(f"[DB ERROR] {e}")


# ══════════════════════════════════════════════════════════════════
#  AUTH HELPERS
# ══════════════════════════════════════════════════════════════════
def hash_password(pw): return hashlib.sha256(pw.encode()).hexdigest()

def get_current_user():
    uid = session.get("user_id")
    if not uid: return None
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE id=%s", (uid,))
        return cur.fetchone()

def get_owner_id():
    user = get_current_user()
    if not user: return None
    return user["owner_id"] if user["role"] == "cashier" else user["id"]

def get_active_branch():
    """Returns branch_id from session, or None (meaning all branches)."""
    return session.get("active_branch_id")

def login_required(f):
    @wraps(f)
    def d(*a, **k):
        if not session.get("user_id"): return redirect("/login")
        return f(*a, **k)
    return d

def api_login_required(f):
    @wraps(f)
    def d(*a, **k):
        if not session.get("user_id"):
            return jsonify({"ok": False, "error": "Not logged in"}), 401
        return f(*a, **k)
    return d

def owner_required(f):
    @wraps(f)
    def d(*a, **k):
        user = get_current_user()
        if not user: return jsonify({"ok": False, "error": "Not logged in"}), 401
        if user["role"] != "owner":
            return jsonify({"ok": False, "error": "Owner access required"}), 403
        return f(*a, **k)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a, **k):
        if not session.get("is_admin"):
            return redirect("/admin/login")
        return f(*a, **k)
    return d


# ══════════════════════════════════════════════════════════════════
#  SUBSCRIPTION HELPERS
# ══════════════════════════════════════════════════════════════════
SUBSCRIPTION_AMOUNT = int(os.getenv("SUBSCRIPTION_AMOUNT", "999"))
TRIAL_DAYS          = int(os.getenv("TRIAL_DAYS", "14"))

def get_or_create_subscription(owner_id):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM subscriptions WHERE owner_id=%s", (owner_id,))
        sub = cur.fetchone()
        if not sub:
            trial_end = datetime.now() + timedelta(days=TRIAL_DAYS)
            cur.execute("""INSERT INTO subscriptions
                (owner_id, status, amount, trial_ends_at, next_billing_date)
                VALUES (%s,'trial',%s,%s,%s) RETURNING *""",
                (owner_id, SUBSCRIPTION_AMOUNT, trial_end,
                 (trial_end + timedelta(days=1)).date()))
            sub = cur.fetchone()
    return dict(sub)

def subscription_status(owner_id):
    sub = get_or_create_subscription(owner_id)
    if sub["status"] == "trial":
        if datetime.now() > sub["trial_ends_at"]:
            with get_db() as conn:
                cur = conn.cursor()
                cur.execute("UPDATE subscriptions SET status='expired' WHERE owner_id=%s", (owner_id,))
            return "expired"
    return sub["status"]

def check_subscription_access(owner_id):
    """Returns True if owner can access the app."""
    status = subscription_status(owner_id)
    return status in ("trial", "active")


# ══════════════════════════════════════════════════════════════════
#  AFRICA'S TALKING
# ══════════════════════════════════════════════════════════════════
AT_USERNAME = os.getenv("AT_USERNAME", "sandbox")
AT_API_KEY  = os.getenv("AT_API_KEY",  "")
AT_ENV      = os.getenv("AT_ENV",      "sandbox")
AT_BASE     = ("https://api.africastalking.com" if AT_ENV == "production"
               else "https://sandbox.africastalking.com")

def format_phone_ke(raw):
    p = str(raw).strip().replace("+","").replace(" ","").replace("-","")
    if p.startswith("0") and len(p) == 10: return "+254" + p[1:]
    if p.startswith("7") and len(p) == 9:  return "+254" + p
    if p.startswith("254") and len(p) == 12: return "+" + p
    return "+" + p.lstrip("0")

def send_sms(phone, message):
    try:
        r = requests.post(f"{AT_BASE}/version1/messaging",
            headers={"apiKey": AT_API_KEY,
                     "Content-Type": "application/x-www-form-urlencoded",
                     "Accept": "application/json"},
            data={"username": AT_USERNAME, "to": format_phone_ke(phone), "message": message},
            timeout=15)
        print(f"[SMS] {phone}: {r.json()}")
        return r.json()
    except Exception as e:
        print(f"[SMS ERROR] {e}")
        return None

def send_otp(phone, purpose):
    code    = str(secrets.randbelow(900000) + 100000)
    expires = datetime.now() + timedelta(minutes=10)
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE otps SET used=TRUE WHERE identifier=%s AND purpose=%s AND used=FALSE",
                    (phone, purpose))
        cur.execute("INSERT INTO otps (identifier,code,purpose,expires_at) VALUES (%s,%s,%s,%s)",
                    (phone, code, purpose, expires))
    msg = f"Your Uniform Shop POS code: {code}. Valid 10 minutes. Do not share."
    if AT_API_KEY:
        send_sms(phone, msg)
    else:
        print(f"[OTP] {phone} → {code}")
    return code

def verify_otp_code(phone, code, purpose):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT * FROM otps WHERE identifier=%s AND code=%s AND purpose=%s
                       AND used=FALSE AND expires_at > NOW() ORDER BY id DESC LIMIT 1""",
                    (phone, code, purpose))
        row = cur.fetchone()
        if not row: return False
        cur.execute("UPDATE otps SET used=TRUE WHERE id=%s", (row["id"],))
        return True


# ══════════════════════════════════════════════════════════════════
#  M-PESA
# ══════════════════════════════════════════════════════════════════
CONSUMER_KEY    = os.getenv("MPESA_CONSUMER_KEY",    "")
CONSUMER_SECRET = os.getenv("MPESA_CONSUMER_SECRET", "")
MPESA_ENV       = os.getenv("MPESA_ENV", "sandbox")
APP_URL         = os.getenv("APP_URL", "https://uniform-shop-pos-1.onrender.com")
CALLBACK_URL    = f"{APP_URL}/mpesa/callback"
SUB_CALLBACK    = f"{APP_URL}/mpesa/subscription-callback"
BASE_URL        = ("https://sandbox.safaricom.co.ke" if MPESA_ENV == "sandbox"
                   else "https://api.safaricom.co.ke")
_SANDBOX_SC     = "174379"
_SANDBOX_PK     = "bfb279f9aa9bdbcf158e97dd71a467cd2e0c893059b10f78e6b72ada1ed2c919"
_token_cache    = {"token": None, "expires": 0}

def mpesa_token():
    now = datetime.now().timestamp()
    if _token_cache["token"] and now < _token_cache["expires"] - 60:
        return _token_cache["token"]
    r = requests.get(f"{BASE_URL}/oauth/v1/generate?grant_type=client_credentials",
                     auth=(CONSUMER_KEY, CONSUMER_SECRET), timeout=15)
    r.raise_for_status()
    data = r.json()
    _token_cache["token"]   = data["access_token"]
    _token_cache["expires"] = now + int(data.get("expires_in", 3600))
    return _token_cache["token"]

def format_phone(raw):
    p = str(raw).strip().replace("+","").replace(" ","").replace("-","")
    if p.startswith("0") and len(p) == 10:   return "254" + p[1:]
    if p.startswith("7") and len(p) == 9:    return "254" + p
    if p.startswith("254") and len(p) == 12: return p
    return "254" + p.lstrip("0")

def resolve_mode(user):
    till    = (user.get("till_number")    or "").strip()
    paybill = (user.get("paybill_number") or "").strip()
    phone   = (user.get("mpesa_phone")    or "").strip()
    passkey = (user.get("mpesa_passkey")  or "").strip()
    if MPESA_ENV == "sandbox":
        return {"shortcode":_SANDBOX_SC,"passkey":_SANDBOX_PK,"party_b":_SANDBOX_SC,
                "tx_type":"CustomerPayBillOnline","acct_ref":None}
    if till:
        return {"shortcode":till,"passkey":passkey or _SANDBOX_PK,
                "party_b":till,"tx_type":"CustomerBuyGoodsOnline","acct_ref":None}
    if paybill:
        return {"shortcode":paybill,"passkey":passkey or _SANDBOX_PK,
                "party_b":paybill,"tx_type":"CustomerPayBillOnline","acct_ref":None}
    if phone:
        proxy = (user.get("mpesa_shortcode") or "").strip() or _SANDBOX_SC
        return {"shortcode":proxy,"passkey":passkey or _SANDBOX_PK,
                "party_b":proxy,"tx_type":"CustomerPayBillOnline",
                "acct_ref":format_phone(phone)}
    raise ValueError("No M-Pesa payment method configured.")

def stk_push(user, customer_phone, amount, product_name, callback_url=None):
    mode      = resolve_mode(user)
    token     = mpesa_token()
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    password  = base64.b64encode(
        f"{mode['shortcode']}{mode['passkey']}{timestamp}".encode()).decode()
    r = requests.post(f"{BASE_URL}/mpesa/stkpush/v1/processrequest",
        json={"BusinessShortCode":mode["shortcode"],"Password":password,
              "Timestamp":timestamp,"TransactionType":mode["tx_type"],
              "Amount":max(1,int(amount)),"PartyA":customer_phone,
              "PartyB":mode["party_b"],"PhoneNumber":customer_phone,
              "CallBackURL":callback_url or CALLBACK_URL,
              "AccountReference":mode["acct_ref"] or product_name[:12],
              "TransactionDesc":"Uniform sale"},
        headers={"Authorization":f"Bearer {token}"}, timeout=15)
    return r.json()

def stk_query(user, ckid):
    mode      = resolve_mode(user)
    token     = mpesa_token()
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    password  = base64.b64encode(
        f"{mode['shortcode']}{mode['passkey']}{timestamp}".encode()).decode()
    r = requests.post(f"{BASE_URL}/mpesa/stkpushquery/v1/query",
        json={"BusinessShortCode":mode["shortcode"],"Password":password,
              "Timestamp":timestamp,"CheckoutRequestID":ckid},
        headers={"Authorization":f"Bearer {token}"}, timeout=15)
    return r.json()

pending = {}
pending_subscriptions = {}


# ══════════════════════════════════════════════════════════════════
#  EMAIL
# ══════════════════════════════════════════════════════════════════
GMAIL_USER = os.getenv("GMAIL_USER", "")
GMAIL_PASS = os.getenv("GMAIL_APP_PASSWORD", "")

def send_email(to_email, subject, body_html, attachment_bytes=None, attachment_name=None):
    if not GMAIL_USER or not GMAIL_PASS:
        print(f"[EMAIL] Not configured — {to_email}: {subject}")
        return False
    try:
        msg = MIMEMultipart("mixed")
        msg["Subject"] = subject
        msg["From"]    = f"Uniform Shop POS <{GMAIL_USER}>"
        msg["To"]      = to_email
        msg.attach(MIMEText(body_html, "html"))
        if attachment_bytes and attachment_name:
            part = MIMEBase("application","octet-stream")
            part.set_payload(attachment_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition",f"attachment; filename={attachment_name}")
            msg.attach(part)
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_PASS)
            s.sendmail(GMAIL_USER, to_email, msg.as_string())
        print(f"[EMAIL] Sent to {to_email}: {subject}")
        return True
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")
        return False


# ══════════════════════════════════════════════════════════════════
#  LOW STOCK ALERTS
# ══════════════════════════════════════════════════════════════════
def check_low_stock_alerts(owner_id, branch_id=None):
    try:
        with get_db() as conn:
            cur = conn.cursor()
            q = """SELECT p.name,p.stock,p.reorder,u.phone,u.business_name
                FROM products p JOIN users u ON u.id=p.user_id
                WHERE p.user_id=%s AND p.stock>0 AND p.stock<=p.reorder"""
            params = [owner_id]
            if branch_id:
                q += " AND (p.branch_id=%s OR p.branch_id IS NULL)"
                params.append(branch_id)
            cur.execute(q, params)
            low = cur.fetchall()
        if low and low[0]["phone"]:
            names = ", ".join(f"{r['name']} ({r['stock']} left)" for r in low)
            send_sms(low[0]["phone"],
                     f"⚠️ {low[0]['business_name']} Low Stock: {names}. Reorder soon!")
    except Exception as e:
        print(f"[STOCK ALERT ERROR] {e}")


# ══════════════════════════════════════════════════════════════════
#  EXPORT HELPERS (openpyxl)
# ══════════════════════════════════════════════════════════════════
def build_excel(headers, rows, sheet_name="Data"):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Header row styling
        header_fill = PatternFill("solid", fgColor="0F1E3D")
        header_font = Font(bold=True, color="F0A500")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    except ImportError:
        return None


# ══════════════════════════════════════════════════════════════════
#  PAGE ROUTES
# ══════════════════════════════════════════════════════════════════
@app.route("/login")
def login_page():
    if session.get("user_id"): return redirect("/")
    return render_template("login.html")

@app.route("/")
@login_required
def index():
    user = get_current_user()
    oid  = get_owner_id()
    if not check_subscription_access(oid):
        return redirect("/subscription")
    return render_template("index.html")

@app.route("/sales")
@login_required
def sales_page(): return render_template("sales.html")

@app.route("/expenses")
@login_required
def expenses_page(): return render_template("expenses.html")

@app.route("/inventory")
@login_required
def inventory_page(): return render_template("inventory.html")

@app.route("/dashboard")
@login_required
def dashboard_page(): return render_template("dashboard.html")

@app.route("/customers")
@login_required
def customers_page(): return render_template("customers.html")

@app.route("/suppliers")
@login_required
def suppliers_page(): return render_template("suppliers.html")

@app.route("/purchase-orders")
@login_required
def purchase_orders_page(): return render_template("purchase_orders.html")

@app.route("/team")
@login_required
def team_page(): return render_template("team.html")

@app.route("/reports")
@login_required
def reports_page(): return render_template("reports.html")

@app.route("/branches")
@login_required
def branches_page(): return render_template("branches.html")

@app.route("/subscription")
@login_required
def subscription_page(): return render_template("subscription.html")

# PWA manifest + service worker
@app.route("/static/manifest.json")
def pwa_manifest():
    user = get_current_user()
    biz  = user["business_name"] if user else "Uniform Shop POS"
    return jsonify({
        "name": biz,
        "short_name": biz[:12],
        "description": "School Uniforms Point of Sale",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#080f20",
        "theme_color": "#0f1e3d",
        "orientation": "portrait",
        "icons": [
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png"}
        ]
    })

@app.route("/sw.js")
def service_worker():
    sw = """
const CACHE='uniform-pos-v1';
const URLS=['/login','/static/manifest.json'];
self.addEventListener('install',e=>e.waitUntil(caches.open(CACHE).then(c=>c.addAll(URLS))));
self.addEventListener('fetch',e=>{
  if(e.request.method!=='GET')return;
  e.respondWith(fetch(e.request).catch(()=>caches.match(e.request)));
});
"""
    return Response(sw, mimetype="application/javascript")


# ══════════════════════════════════════════════════════════════════
#  ADMIN PANEL
# ══════════════════════════════════════════════════════════════════
ADMIN_EMAIL    = os.getenv("ADMIN_EMAIL",    "")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")

@app.route("/admin")
def admin_redirect(): return redirect("/admin/login")

@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    if request.method == "GET":
        return render_template("admin_login.html")
    d    = request.json or {}
    email = d.get("email","").strip().lower()
    pwd   = d.get("password","").strip()
    if email == ADMIN_EMAIL.lower() and pwd == ADMIN_PASSWORD:
        session["is_admin"] = True
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Invalid admin credentials"}), 401

@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("is_admin", None)
    return jsonify({"ok": True})

@app.route("/admin/dashboard")
@admin_required
def admin_dashboard(): return render_template("admin.html")

@app.route("/api/admin/stats")
@admin_required
def admin_stats():
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) as c FROM users WHERE role='owner'")
            total_owners = cur.fetchone()["c"]
            cur.execute("SELECT COUNT(*) as c FROM users WHERE role='cashier'")
            total_cashiers = cur.fetchone()["c"]
            cur.execute("SELECT COUNT(*) as c FROM sales")
            total_sales = cur.fetchone()["c"]
            cur.execute("SELECT COALESCE(SUM(total_amount),0) as v FROM sales")
            total_revenue = cur.fetchone()["v"]
            cur.execute("SELECT COUNT(*) as c FROM subscriptions WHERE status='active'")
            active_subs = cur.fetchone()["c"]
            cur.execute("SELECT COUNT(*) as c FROM subscriptions WHERE status='trial'")
            trial_subs = cur.fetchone()["c"]
            cur.execute("""SELECT COUNT(*) as c FROM subscriptions
                WHERE status='active' AND last_paid_at IS NOT NULL""")
            paid_subs = cur.fetchone()["c"]
            platform_revenue = paid_subs * SUBSCRIPTION_AMOUNT
            cur.execute("""SELECT u.id,u.business_name,u.name,u.email,u.phone,
                u.created_at,u.is_active,
                s.status as sub_status,s.trial_ends_at,s.last_paid_at,
                (SELECT COUNT(*) FROM sales WHERE user_id=u.id) as sale_count,
                (SELECT COUNT(*) FROM products WHERE user_id=u.id) as product_count,
                (SELECT COUNT(*) FROM branches WHERE owner_id=u.id) as branch_count
                FROM users u
                LEFT JOIN subscriptions s ON s.owner_id=u.id
                WHERE u.role='owner'
                ORDER BY u.created_at DESC""")
            owners = [dict(r) for r in cur.fetchall()]
        return jsonify({"ok":True,"stats":{
            "total_owners":total_owners,"total_cashiers":total_cashiers,
            "total_sales":total_sales,"total_revenue":total_revenue,
            "active_subs":active_subs,"trial_subs":trial_subs,
            "platform_revenue":platform_revenue,"owners":owners,
        }})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500

@app.route("/api/admin/suspend/<int:uid>", methods=["POST"])
@admin_required
def admin_suspend(uid):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE users SET is_active=FALSE WHERE id=%s AND role='owner'", (uid,))
        cur.execute("UPDATE subscriptions SET status='suspended' WHERE owner_id=%s", (uid,))
    return jsonify({"ok":True})

@app.route("/api/admin/reactivate/<int:uid>", methods=["POST"])
@admin_required
def admin_reactivate(uid):
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE users SET is_active=TRUE WHERE id=%s AND role='owner'", (uid,))
        cur.execute("UPDATE subscriptions SET status='active' WHERE owner_id=%s", (uid,))
    return jsonify({"ok":True})


# ══════════════════════════════════════════════════════════════════
#  AUTH API
# ══════════════════════════════════════════════════════════════════
@app.route("/api/signup", methods=["POST"])
def api_signup():
    d        = request.json or {}
    email    = (d.get("email")         or "").strip().lower()
    password = (d.get("password")      or "").strip()
    name     = (d.get("name")          or "").strip()
    business = (d.get("business_name") or "").strip()
    phone    = (d.get("phone")         or "").strip()
    if not all([email, password, name, business]):
        return jsonify({"ok":False,"error":"Please fill in all required fields"}), 400
    if len(password) < 6:
        return jsonify({"ok":False,"error":"Password must be at least 6 characters"}), 400
    mpesa_phone    = (d.get("mpesa_phone")    or "").strip()
    till_number    = (d.get("till_number")    or "").strip()
    paybill_number = (d.get("paybill_number") or "").strip()
    if not any([mpesa_phone, till_number, paybill_number]):
        return jsonify({"ok":False,"error":"Please provide at least one M-Pesa payment method"}), 400
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("""INSERT INTO users
                (email,password_hash,name,business_name,phone,mpesa_phone,till_number,paybill_number)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (email,hash_password(password),name,business,
                 phone,mpesa_phone,till_number,paybill_number))
            uid = cur.fetchone()["id"]
        session["user_id"] = uid
        get_or_create_subscription(uid)
        if AT_API_KEY and phone:
            send_otp(phone, "signup")
            return jsonify({"ok":True,"otp_required":True,"phone":format_phone_ke(phone)})
        return jsonify({"ok":True,"otp_required":False})
    except psycopg2.errors.UniqueViolation:
        return jsonify({"ok":False,"error":"An account with this email already exists"}), 400
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500

@app.route("/api/login", methods=["POST"])
def api_login():
    d        = request.json or {}
    email    = (d.get("email")    or "").strip().lower()
    password = (d.get("password") or "").strip()
    if not email or not password:
        return jsonify({"ok":False,"error":"Please enter email and password"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE email=%s", (email,))
        user = cur.fetchone()
    if not user or user["password_hash"] != hash_password(password):
        return jsonify({"ok":False,"error":"Incorrect email or password"}), 401
    if not user.get("is_active", True):
        return jsonify({"ok":False,"error":"This account has been suspended. Contact support."}), 403
    if user["role"] == "cashier":
        session["user_id"] = user["id"]
        if user.get("branch_id"):
            session["active_branch_id"] = user["branch_id"]
        return jsonify({"ok":True,"otp_required":False})
    session["_pending_uid"] = user["id"]
    phone = user.get("phone","")
    if AT_API_KEY and phone:
        send_otp(phone, "login")
        return jsonify({"ok":True,"otp_required":True,"phone":format_phone_ke(phone)})
    session["user_id"] = user["id"]
    session.pop("_pending_uid", None)
    return jsonify({"ok":True,"otp_required":False})

@app.route("/api/verify-otp", methods=["POST"])
def api_verify_otp():
    d       = request.json or {}
    phone   = (d.get("phone")   or "").strip()
    code    = (d.get("code")    or "").strip()
    purpose = (d.get("purpose") or "login").strip()
    pn = phone.replace("+","").replace(" ","")
    if pn.startswith("254"): pn = "0" + pn[3:]
    if not (verify_otp_code(pn, code, purpose) or
            verify_otp_code(format_phone_ke(pn), code, purpose) or
            verify_otp_code(phone, code, purpose)):
        return jsonify({"ok":False,"error":"Invalid or expired code."}), 400
    if purpose in ("login", "signup"):
        uid = session.pop("_pending_uid", session.get("user_id"))
        if uid: session["user_id"] = uid
    return jsonify({"ok":True})

@app.route("/api/resend-otp", methods=["POST"])
def api_resend_otp():
    d = request.json or {}
    send_otp((d.get("phone") or "").strip(), (d.get("purpose") or "login"))
    return jsonify({"ok":True})

@app.route("/api/forgot-password", methods=["POST"])
def api_forgot_password():
    phone = (request.json or {}).get("phone","").strip()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id FROM users WHERE phone=%s OR phone=%s",
                    (phone, format_phone_ke(phone)))
        row = cur.fetchone()
    if not row:
        return jsonify({"ok":False,"error":"No account found with that phone number"}), 404
    send_otp(phone, "reset")
    return jsonify({"ok":True})

@app.route("/api/reset-password", methods=["POST"])
def api_reset_password():
    d       = request.json or {}
    phone   = (d.get("phone")         or "").strip()
    new_pwd = (d.get("new_password")  or "").strip()
    if len(new_pwd) < 6:
        return jsonify({"ok":False,"error":"Password must be at least 6 characters"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE users SET password_hash=%s WHERE phone=%s OR phone=%s",
                    (hash_password(new_pwd), phone, format_phone_ke(phone)))
    return jsonify({"ok":True})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok":True})

@app.route("/api/me")
@api_login_required
def api_me():
    user = get_current_user()
    oid  = get_owner_id()
    sub  = get_or_create_subscription(oid)
    mode = ("Till"    if user.get("till_number")    else
            "Paybill" if user.get("paybill_number") else
            "Phone"   if user.get("mpesa_phone")    else "Not set")
    return jsonify({"ok":True,"user":{
        "id":user["id"],"name":user["name"],
        "business_name":user["business_name"],
        "email":user["email"],"phone":user.get("phone",""),
        "role":user["role"],"mpesa_mode":mode,
        "branch_id":user.get("branch_id"),
        "subscription":{"status":sub["status"],
                        "trial_ends_at":str(sub.get("trial_ends_at","") or ""),
                        "next_billing_date":str(sub.get("next_billing_date","") or "")},
    }})


# ══════════════════════════════════════════════════════════════════
#  BRANCHES
# ══════════════════════════════════════════════════════════════════
@app.route("/api/branches")
@api_login_required
def get_branches():
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT b.*,
            (SELECT COUNT(*) FROM sales WHERE branch_id=b.id) as sale_count,
            (SELECT COUNT(*) FROM products WHERE branch_id=b.id) as product_count,
            (SELECT COUNT(*) FROM users WHERE branch_id=b.id AND role='cashier') as cashier_count
            FROM branches b WHERE b.owner_id=%s ORDER BY b.created_at""", (oid,))
        rows = cur.fetchall()
    return jsonify({"ok":True,"branches":[dict(r) for r in rows]})

@app.route("/api/branches", methods=["POST"])
@owner_required
def create_branch():
    oid  = get_owner_id()
    d    = request.json or {}
    name = (d.get("name") or "").strip()
    if not name: return jsonify({"ok":False,"error":"Branch name required"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO branches (owner_id,name,location,inventory_mode)
            VALUES (%s,%s,%s,%s) RETURNING *""",
            (oid, name, d.get("location","").strip(),
             d.get("inventory_mode","shared")))
        row = dict(cur.fetchone())
    return jsonify({"ok":True,"branch":row})

@app.route("/api/branches/<int:bid>", methods=["PUT"])
@owner_required
def update_branch(bid):
    oid = get_owner_id()
    d   = request.json or {}
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""UPDATE branches SET name=%s,location=%s,inventory_mode=%s,is_active=%s
            WHERE id=%s AND owner_id=%s RETURNING *""",
            (d.get("name"),d.get("location",""),
             d.get("inventory_mode","shared"),
             d.get("is_active",True), bid, oid))
        row = cur.fetchone()
    if not row: return jsonify({"ok":False,"error":"Branch not found"}), 404
    return jsonify({"ok":True,"branch":dict(row)})

@app.route("/api/branches/<int:bid>", methods=["DELETE"])
@owner_required
def delete_branch(bid):
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM branches WHERE id=%s AND owner_id=%s", (bid,oid))
    return jsonify({"ok":True})

@app.route("/api/branches/switch", methods=["POST"])
@api_login_required
def switch_branch():
    bid = (request.json or {}).get("branch_id")
    if bid is None:
        session.pop("active_branch_id", None)
    else:
        session["active_branch_id"] = int(bid)
    return jsonify({"ok":True,"active_branch_id":bid})

@app.route("/api/branches/combined-stats")
@api_login_required
def combined_branch_stats():
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT b.id,b.name,b.location,b.inventory_mode,
            COALESCE((SELECT SUM(total_amount) FROM sales WHERE branch_id=b.id AND DATE(created_at)=CURRENT_DATE),0) as today_revenue,
            COALESCE((SELECT COUNT(*) FROM sales WHERE branch_id=b.id AND DATE(created_at)=CURRENT_DATE),0) as today_txns,
            COALESCE((SELECT SUM(total_amount) FROM sales WHERE branch_id=b.id AND DATE_TRUNC('month',created_at)=DATE_TRUNC('month',NOW())),0) as month_revenue,
            COALESCE((SELECT COUNT(*) FROM products WHERE branch_id=b.id),0) as products,
            COALESCE((SELECT COUNT(*) FROM products WHERE branch_id=b.id AND stock<=0),0) as out_of_stock
            FROM branches b WHERE b.owner_id=%s AND b.is_active=TRUE ORDER BY b.name""", (oid,))
        rows = cur.fetchall()
    return jsonify({"ok":True,"branches":[dict(r) for r in rows]})


# ══════════════════════════════════════════════════════════════════
#  PRODUCTS
# ══════════════════════════════════════════════════════════════════
@app.route("/api/products")
@api_login_required
def get_products():
    oid = get_owner_id()
    bid = get_active_branch()
    with get_db() as conn:
        cur = conn.cursor()
        if bid:
            # Get branch inventory mode
            cur.execute("SELECT inventory_mode FROM branches WHERE id=%s AND owner_id=%s",
                        (bid, oid))
            br = cur.fetchone()
            if br and br["inventory_mode"] == "independent":
                cur.execute("SELECT * FROM products WHERE user_id=%s AND branch_id=%s ORDER BY category,name",
                            (oid, bid))
            else:
                # shared mode — central products (no branch_id)
                cur.execute("SELECT * FROM products WHERE user_id=%s AND branch_id IS NULL ORDER BY category,name",
                            (oid,))
        else:
            cur.execute("SELECT * FROM products WHERE user_id=%s ORDER BY category,name", (oid,))
        rows = cur.fetchall()
    return jsonify({"ok":True,"products":[dict(r) for r in rows]})

@app.route("/api/products", methods=["POST"])
@api_login_required
def add_product():
    oid = get_owner_id()
    d   = request.json or {}
    if not (d.get("name") or "").strip():
        return jsonify({"ok":False,"error":"Product name is required"}), 400
    branch_id = d.get("branch_id") or get_active_branch()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO products
            (user_id,branch_id,name,category,barcode,price,wholesale_price,bulk_price,
             discount_pct,discount_fixed,vat_exempt,stock,reorder)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (oid, branch_id, d["name"].strip(), d.get("category","General"),
             d.get("barcode","").strip(), int(d.get("price",0)),
             int(d.get("wholesale_price",0)), int(d.get("bulk_price",0)),
             int(d.get("discount_pct",0)), int(d.get("discount_fixed",0)),
             bool(d.get("vat_exempt",False)),
             int(d.get("stock",0)), int(d.get("reorder",5))))
        row = dict(cur.fetchone())
    return jsonify({"ok":True,"product":row})

@app.route("/api/products/<int:pid>", methods=["PUT"])
@api_login_required
def update_product(pid):
    oid = get_owner_id()
    d   = request.json or {}
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""UPDATE products SET
            name=%s,category=%s,barcode=%s,price=%s,wholesale_price=%s,bulk_price=%s,
            discount_pct=%s,discount_fixed=%s,vat_exempt=%s,stock=%s,reorder=%s
            WHERE id=%s AND user_id=%s RETURNING *""",
            (d.get("name"),d.get("category","General"),d.get("barcode",""),
             int(d.get("price",0)),int(d.get("wholesale_price",0)),
             int(d.get("bulk_price",0)),int(d.get("discount_pct",0)),
             int(d.get("discount_fixed",0)),bool(d.get("vat_exempt",False)),
             int(d.get("stock",0)),int(d.get("reorder",5)),pid,oid))
        row = cur.fetchone()
    if not row: return jsonify({"ok":False,"error":"Product not found"}), 404
    return jsonify({"ok":True,"product":dict(row)})

@app.route("/api/products/<int:pid>", methods=["DELETE"])
@api_login_required
def delete_product(pid):
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM products WHERE id=%s AND user_id=%s", (pid,oid))
    return jsonify({"ok":True})


# ══════════════════════════════════════════════════════════════════
#  CUSTOMERS
# ══════════════════════════════════════════════════════════════════
@app.route("/api/customers")
@api_login_required
def get_customers():
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE owner_id=%s ORDER BY name", (oid,))
        rows = cur.fetchall()
    return jsonify({"ok":True,"customers":[dict(r) for r in rows]})

@app.route("/api/customers", methods=["POST"])
@api_login_required
def add_customer():
    oid  = get_owner_id()
    d    = request.json or {}
    name = (d.get("name") or "").strip()
    if not name: return jsonify({"ok":False,"error":"Customer name required"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO customers (owner_id,name,phone,email,customer_type)
            VALUES (%s,%s,%s,%s,%s) RETURNING *""",
            (oid,name,d.get("phone","").strip(),
             d.get("email","").strip(),d.get("customer_type","retail")))
        row = dict(cur.fetchone())
    return jsonify({"ok":True,"customer":row})

@app.route("/api/customers/<int:cid>", methods=["PUT"])
@api_login_required
def update_customer(cid):
    oid = get_owner_id()
    d   = request.json or {}
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""UPDATE customers SET name=%s,phone=%s,email=%s,customer_type=%s
            WHERE id=%s AND owner_id=%s RETURNING *""",
            (d.get("name"),d.get("phone",""),d.get("email",""),
             d.get("customer_type","retail"),cid,oid))
        row = cur.fetchone()
    if not row: return jsonify({"ok":False,"error":"Customer not found"}), 404
    return jsonify({"ok":True,"customer":dict(row)})

@app.route("/api/customers/<int:cid>", methods=["DELETE"])
@api_login_required
def delete_customer(cid):
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM customers WHERE id=%s AND owner_id=%s", (cid,oid))
    return jsonify({"ok":True})

@app.route("/api/customers/<int:cid>/history")
@api_login_required
def customer_history(cid):
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT s.txn_id,s.total_amount,s.created_at,s.mpesa_code
            FROM sales s WHERE s.customer_id=%s AND s.user_id=%s
            ORDER BY s.created_at DESC LIMIT 20""", (cid,oid))
        rows = cur.fetchall()
    return jsonify({"ok":True,"sales":[dict(r) for r in rows]})


# ══════════════════════════════════════════════════════════════════
#  SUPPLIERS
# ══════════════════════════════════════════════════════════════════
@app.route("/api/suppliers")
@api_login_required
def get_suppliers():
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM suppliers WHERE owner_id=%s ORDER BY name", (oid,))
        rows = cur.fetchall()
    return jsonify({"ok":True,"suppliers":[dict(r) for r in rows]})

@app.route("/api/suppliers", methods=["POST"])
@api_login_required
def add_supplier():
    oid  = get_owner_id()
    d    = request.json or {}
    name = (d.get("name") or "").strip()
    if not name: return jsonify({"ok":False,"error":"Supplier name required"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO suppliers (owner_id,name,phone,email,address,balance_owed)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *""",
            (oid,name,d.get("phone",""),d.get("email",""),
             d.get("address",""),int(d.get("balance_owed",0))))
        row = dict(cur.fetchone())
    return jsonify({"ok":True,"supplier":row})

@app.route("/api/suppliers/<int:sid>", methods=["PUT"])
@api_login_required
def update_supplier(sid):
    oid = get_owner_id()
    d   = request.json or {}
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""UPDATE suppliers SET name=%s,phone=%s,email=%s,address=%s,balance_owed=%s
            WHERE id=%s AND owner_id=%s RETURNING *""",
            (d.get("name"),d.get("phone",""),d.get("email",""),
             d.get("address",""),int(d.get("balance_owed",0)),sid,oid))
        row = cur.fetchone()
    if not row: return jsonify({"ok":False,"error":"Supplier not found"}), 404
    return jsonify({"ok":True,"supplier":dict(row)})

@app.route("/api/suppliers/<int:sid>/payment", methods=["POST"])
@api_login_required
def record_supplier_payment(sid):
    oid = get_owner_id()
    d   = request.json or {}
    amt = int(d.get("amount",0))
    if amt <= 0: return jsonify({"ok":False,"error":"Invalid amount"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""UPDATE suppliers SET balance_owed=GREATEST(balance_owed-%s,0)
            WHERE id=%s AND owner_id=%s RETURNING *""", (amt,sid,oid))
        row = cur.fetchone()
    if not row: return jsonify({"ok":False,"error":"Supplier not found"}), 404
    return jsonify({"ok":True,"supplier":dict(row)})

@app.route("/api/suppliers/<int:sid>", methods=["DELETE"])
@api_login_required
def delete_supplier(sid):
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM suppliers WHERE id=%s AND owner_id=%s", (sid,oid))
    return jsonify({"ok":True})


# ══════════════════════════════════════════════════════════════════
#  PURCHASE ORDERS
# ══════════════════════════════════════════════════════════════════
@app.route("/api/purchase-orders")
@api_login_required
def get_purchase_orders():
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT po.*,s.name as supplier_name,b.name as branch_name,
            json_agg(json_build_object('name',poi.name,'qty',poi.qty,
                'unit_cost',poi.unit_cost,'subtotal',poi.subtotal)) as items
            FROM purchase_orders po
            LEFT JOIN suppliers s ON s.id=po.supplier_id
            LEFT JOIN branches b ON b.id=po.branch_id
            LEFT JOIN purchase_order_items poi ON poi.order_id=po.id
            WHERE po.owner_id=%s GROUP BY po.id,s.name,b.name ORDER BY po.created_at DESC""", (oid,))
        rows = cur.fetchall()
    return jsonify({"ok":True,"orders":[dict(r) for r in rows]})

@app.route("/api/purchase-orders", methods=["POST"])
@api_login_required
def create_purchase_order():
    oid   = get_owner_id()
    d     = request.json or {}
    items = d.get("items",[])
    if not items: return jsonify({"ok":False,"error":"No items in order"}), 400
    bid = d.get("branch_id") or get_active_branch()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO purchase_orders (owner_id,branch_id,supplier_id,total_cost,note)
            VALUES (%s,%s,%s,%s,%s) RETURNING id""",
            (oid,bid,d.get("supplier_id"),int(d.get("total_cost",0)),d.get("note","")))
        order_id = cur.fetchone()["id"]
        for item in items:
            cur.execute("""INSERT INTO purchase_order_items
                (order_id,product_id,name,qty,unit_cost,subtotal)
                VALUES (%s,%s,%s,%s,%s,%s)""",
                (order_id,item.get("product_id"),item["name"],
                 int(item["qty"]),int(item["unit_cost"]),int(item["subtotal"])))
            if item.get("product_id"):
                cur.execute("UPDATE products SET stock=stock+%s WHERE id=%s AND user_id=%s",
                            (int(item["qty"]),item["product_id"],oid))
    return jsonify({"ok":True,"order_id":order_id})


# ══════════════════════════════════════════════════════════════════
#  PAYMENT
# ══════════════════════════════════════════════════════════════════
@app.route("/api/pay", methods=["POST"])
@api_login_required
def pay():
    user = get_current_user()
    oid  = get_owner_id()
    try:
        d        = request.json
        phone    = format_phone(d["phone"])
        customer = d.get("customer") or "Walk-in"
        items    = d.get("items",[])
        amount   = int(d.get("total") or sum(i.get("subtotal",0) for i in items))
        first    = items[0]["product"]["name"] if items else "Uniforms"
        bid      = d.get("branch_id") or get_active_branch()
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM users WHERE id=%s", (oid,))
            owner = cur.fetchone()
        resp = stk_push(owner, phone, amount, first)
        if str(resp.get("ResponseCode")) == "0":
            ckid = resp["CheckoutRequestID"]
            pending[ckid] = {
                "user_id":user["id"],"owner_id":oid,"phone":phone,
                "customer":customer,"customer_id":d.get("customer_id"),
                "branch_id":bid,"items":items,"amount":amount,
                "timestamp":datetime.now(),
                "vat_amount":int(d.get("vat_amount",0)),
                "discount":int(d.get("discount",0)),
                "price_tier":d.get("price_tier","retail"),
                "status":"pending",
            }
            return jsonify({"ok":True,"checkout_id":ckid})
        err = (resp.get("errorMessage") or resp.get("ResponseDescription")
               or resp.get("ResultDesc") or "M-Pesa request failed.")
        return jsonify({"ok":False,"error":err}), 400
    except ValueError as e: return jsonify({"ok":False,"error":str(e)}), 400
    except Exception as e:
        print(f"[PAY ERROR] {e}")
        return jsonify({"ok":False,"error":str(e)}), 500

@app.route("/mpesa/callback", methods=["POST"])
def mpesa_callback():
    body = request.json or {}
    stk  = body.get("Body",{}).get("stkCallback",{})
    ckid = stk.get("CheckoutRequestID")
    code = stk.get("ResultCode")
    if ckid in pending:
        entry = pending[ckid]
        if code == 0:
            items      = stk.get("CallbackMetadata",{}).get("Item",[])
            mpesa_code = next((i["Value"] for i in items if i["Name"]=="MpesaReceiptNumber"),"N/A")
            entry["status"]     = "completed"
            entry["mpesa_code"] = mpesa_code
            sid = _record_sale(entry, mpesa_code)
            entry["sale_id"] = sid
        else:
            entry["status"] = "failed"
    return jsonify({"ResultCode":0,"ResultDesc":"Accepted"})

@app.route("/mpesa/subscription-callback", methods=["POST"])
def subscription_callback():
    body = request.json or {}
    stk  = body.get("Body",{}).get("stkCallback",{})
    ckid = stk.get("CheckoutRequestID")
    code = stk.get("ResultCode")
    if ckid in pending_subscriptions:
        entry = pending_subscriptions[ckid]
        oid   = entry["owner_id"]
        if code == 0:
            next_date = (date.today() + timedelta(days=30))
            with get_db() as conn:
                cur = conn.cursor()
                cur.execute("""UPDATE subscriptions SET status='active',
                    last_paid_at=NOW(), next_billing_date=%s
                    WHERE owner_id=%s""", (next_date, oid))
            pending_subscriptions.pop(ckid, None)
        else:
            pending_subscriptions.pop(ckid, None)
    return jsonify({"ResultCode":0,"ResultDesc":"Accepted"})

@app.route("/api/status/<ckid>")
@api_login_required
def check_status(ckid):
    entry = pending.get(ckid)
    if not entry: return jsonify({"status":"completed"})
    if entry["status"] == "completed":
        return jsonify({"status":"completed","sale_id":entry.get("sale_id"),
                        "mpesa_code":entry.get("mpesa_code","")})
    if entry["status"] == "failed":
        pending.pop(ckid,None)
        return jsonify({"status":"failed","reason":"Payment cancelled or failed"})
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM users WHERE id=%s", (entry["owner_id"],))
            owner = cur.fetchone()
        if owner:
            q    = stk_query(owner, ckid)
            code = q.get("ResultCode")
            if code == 0:
                mc = q.get("MpesaReceiptNumber","QUERY-OK")
                entry["status"]     = "completed"
                entry["mpesa_code"] = mc
                sid = _record_sale(entry, mc)
                entry["sale_id"] = sid
                return jsonify({"status":"completed","sale_id":sid,"mpesa_code":mc})
            if code in (1032,1,17,2001):
                entry["status"] = "failed"
                pending.pop(ckid,None)
                return jsonify({"status":"failed","reason":q.get("ResultDesc","Payment failed")})
    except Exception as e: print(f"[QUERY ERROR] {e}")
    return jsonify({"status":"pending"})

def _record_sale(entry, mpesa_code):
    try:
        oid = entry["owner_id"]
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) as c FROM sales WHERE user_id=%s", (oid,))
            n      = cur.fetchone()["c"] + 1
            txn_id = f"TXN-{n:04d}"
            ts     = entry["timestamp"]
            cur.execute("""INSERT INTO sales
                (user_id,cashier_id,customer_id,branch_id,txn_id,customer,mpesa_code,
                 total_amount,vat_amount,discount,price_tier,phone,created_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (oid,entry["user_id"],entry.get("customer_id"),entry.get("branch_id"),
                 txn_id,entry["customer"],mpesa_code,entry["amount"],
                 entry.get("vat_amount",0),entry.get("discount",0),
                 entry.get("price_tier","retail"),entry["phone"],ts))
            sale_id = cur.fetchone()["id"]
            for item in entry.get("items",[]):
                p   = item["product"]
                qty = int(item["qty"])
                cur.execute("""INSERT INTO sale_items
                    (sale_id,product_id,name,category,price,qty,subtotal)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                    (sale_id,p.get("id"),p["name"],p.get("category",""),
                     int(item.get("unit_price",p["price"])),qty,int(item.get("subtotal",0))))
                if p.get("id"):
                    cur.execute("UPDATE products SET stock=GREATEST(stock-%s,0) WHERE id=%s AND user_id=%s",
                                (qty,p["id"],oid))
            if entry.get("customer_id"):
                pts = max(1, int(entry["amount"]) // 100)
                cur.execute("""UPDATE customers SET loyalty_points=loyalty_points+%s,
                    total_spent=total_spent+%s WHERE id=%s AND owner_id=%s""",
                    (pts,entry["amount"],entry["customer_id"],oid))
        print(f"[SALE] {txn_id} KES {entry['amount']} | {mpesa_code}")
        check_low_stock_alerts(oid, entry.get("branch_id"))
        return sale_id
    except Exception as e:
        print(f"[SALE ERROR] {e}")
        return None


# ══════════════════════════════════════════════════════════════════
#  SUBSCRIPTION API
# ══════════════════════════════════════════════════════════════════
@app.route("/api/subscription")
@api_login_required
def get_subscription():
    oid = get_owner_id()
    sub = get_or_create_subscription(oid)
    return jsonify({"ok":True,"subscription":dict(sub)})

@app.route("/api/subscription/pay", methods=["POST"])
@api_login_required
def pay_subscription():
    user = get_current_user()
    oid  = get_owner_id()
    phone = format_phone(user.get("phone","") or user.get("mpesa_phone",""))
    if not phone:
        return jsonify({"ok":False,"error":"No phone number on account for M-Pesa"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE id=%s", (oid,))
        owner = cur.fetchone()
    try:
        resp = stk_push(owner, phone, SUBSCRIPTION_AMOUNT,
                        "Subscription", SUB_CALLBACK)
        if str(resp.get("ResponseCode")) == "0":
            ckid = resp["CheckoutRequestID"]
            pending_subscriptions[ckid] = {"owner_id":oid}
            with get_db() as conn:
                cur = conn.cursor()
                cur.execute("UPDATE subscriptions SET checkout_id=%s WHERE owner_id=%s",
                            (ckid, oid))
            return jsonify({"ok":True,"checkout_id":ckid})
        err = resp.get("errorMessage") or resp.get("ResultDesc") or "M-Pesa request failed."
        return jsonify({"ok":False,"error":err}), 400
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500

@app.route("/api/subscription/status/<ckid>")
@api_login_required
def subscription_payment_status(ckid):
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT status FROM subscriptions WHERE owner_id=%s", (oid,))
        row = cur.fetchone()
    status = row["status"] if row else "unknown"
    if status == "active":
        return jsonify({"status":"completed"})
    if ckid in pending_subscriptions:
        return jsonify({"status":"pending"})
    return jsonify({"status":"pending"})


# ══════════════════════════════════════════════════════════════════
#  REFUNDS
# ══════════════════════════════════════════════════════════════════
@app.route("/api/refunds", methods=["POST"])
@api_login_required
def process_refund():
    oid     = get_owner_id()
    d       = request.json or {}
    sale_id = d.get("sale_id")
    reason  = d.get("reason","Customer return")
    note    = d.get("note","")
    if not sale_id: return jsonify({"ok":False,"error":"Sale ID required"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM sales WHERE id=%s AND user_id=%s", (sale_id,oid))
        sale = cur.fetchone()
        if not sale: return jsonify({"ok":False,"error":"Sale not found"}), 404
        cur.execute("""INSERT INTO refunds (owner_id,sale_id,txn_id,reason,note,total_amount)
            VALUES (%s,%s,%s,%s,%s,%s)""",
            (oid,sale_id,sale["txn_id"],reason,note,sale["total_amount"]))
        cur.execute("SELECT * FROM sale_items WHERE sale_id=%s", (sale_id,))
        items = cur.fetchall()
        for item in items:
            if item["product_id"]:
                cur.execute("UPDATE products SET stock=stock+%s WHERE id=%s AND user_id=%s",
                            (item["qty"],item["product_id"],oid))
    return jsonify({"ok":True})


# ══════════════════════════════════════════════════════════════════
#  SALES
# ══════════════════════════════════════════════════════════════════
@app.route("/api/sales")
@api_login_required
def get_sales():
    oid    = get_owner_id()
    bid    = request.args.get("branch_id") or get_active_branch()
    period = request.args.get("period","daily")
    base   = """SELECT s.*,
        u.name as cashier_name,
        b.name as branch_name,
        json_agg(json_build_object('name',si.name,'category',si.category,
            'qty',si.qty,'price',si.price,'subtotal',si.subtotal)) AS items
        FROM sales s
        LEFT JOIN users u ON u.id=s.cashier_id
        LEFT JOIN branches b ON b.id=s.branch_id
        LEFT JOIN sale_items si ON si.sale_id=s.id
        WHERE s.user_id=%s"""
    params = [oid]
    if bid:
        base += " AND s.branch_id=%s"; params.append(bid)
    if period == "daily":
        base += " AND DATE(s.created_at)=%s"
        params.append(request.args.get("date",date.today().isoformat()))
    elif period == "monthly":
        base += " AND EXTRACT(YEAR FROM s.created_at)=%s AND EXTRACT(MONTH FROM s.created_at)=%s"
        params += [request.args.get("year",date.today().year),
                   request.args.get("month",date.today().month)]
    elif period == "yearly":
        base += " AND EXTRACT(YEAR FROM s.created_at)=%s"
        params.append(request.args.get("year",date.today().year))
    elif period == "range":
        base += " AND DATE(s.created_at) BETWEEN %s AND %s"
        params += [request.args.get("from",date.today().isoformat()),
                   request.args.get("to",date.today().isoformat())]
    base += " GROUP BY s.id,u.name,b.name ORDER BY s.created_at DESC"
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(base, params)
        rows = cur.fetchall()
    sales   = [dict(r) for r in rows]
    summary = {
        "total_revenue":     sum(r["total_amount"] for r in rows),
        "total_transactions": len(rows),
        "total_vat":         sum(r.get("vat_amount",0) or 0 for r in rows),
        "total_discounts":   sum(r.get("discount",0) or 0 for r in rows),
    }
    return jsonify({"ok":True,"sales":sales,"summary":summary})


# ══════════════════════════════════════════════════════════════════
#  EXPENSES
# ══════════════════════════════════════════════════════════════════
@app.route("/api/expenses")
@api_login_required
def get_expenses():
    oid    = get_owner_id()
    bid    = request.args.get("branch_id") or get_active_branch()
    period = request.args.get("period","daily")
    q      = "SELECT * FROM expenses WHERE user_id=%s"
    params = [oid]
    if bid: q += " AND (branch_id=%s OR branch_id IS NULL)"; params.append(bid)
    if period == "daily":
        q += " AND DATE(created_at)=%s"
        params.append(request.args.get("date",date.today().isoformat()))
    elif period == "monthly":
        q += " AND EXTRACT(YEAR FROM created_at)=%s AND EXTRACT(MONTH FROM created_at)=%s"
        params += [request.args.get("year",date.today().year),
                   request.args.get("month",date.today().month)]
    elif period == "yearly":
        q += " AND EXTRACT(YEAR FROM created_at)=%s"
        params.append(request.args.get("year",date.today().year))
    q += " ORDER BY created_at DESC"
    with get_db() as conn:
        cur = conn.cursor(); cur.execute(q,params); rows = cur.fetchall()
    return jsonify({"ok":True,"expenses":[dict(r) for r in rows],
                    "summary":{"total":sum(r["amount"] for r in rows),"count":len(rows)}})

@app.route("/api/expenses", methods=["POST"])
@api_login_required
def add_expense():
    oid = get_owner_id()
    d   = request.json or {}
    desc = (d.get("description") or "").strip()
    amt  = int(d.get("amount",0))
    if not desc or amt <= 0: return jsonify({"ok":False,"error":"Description and amount required"}), 400
    bid = d.get("branch_id") or get_active_branch()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""INSERT INTO expenses (user_id,branch_id,description,amount,category,note)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING *""",
            (oid,bid,desc,amt,d.get("category","General"),d.get("note","")))
        row = dict(cur.fetchone())
    return jsonify({"ok":True,"expense":row})

@app.route("/api/expenses/<int:eid>", methods=["DELETE"])
@api_login_required
def delete_expense(eid):
    oid = get_owner_id()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM expenses WHERE id=%s AND user_id=%s", (eid,oid))
    return jsonify({"ok":True})


# ══════════════════════════════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════════════════════════════
@app.route("/api/dashboard")
@api_login_required
def get_dashboard():
    oid    = get_owner_id()
    period = request.args.get("period","daily")
    bid    = request.args.get("branch_id") or get_active_branch()
    if period == "daily":
        sf = "DATE(created_at)=CURRENT_DATE"
        cf = "created_at >= NOW() - INTERVAL '7 days'"
    elif period == "weekly":
        sf = "created_at >= DATE_TRUNC('week',NOW())"
        cf = "created_at >= NOW() - INTERVAL '7 days'"
    elif period == "monthly":
        sf = "DATE_TRUNC('month',created_at)=DATE_TRUNC('month',NOW())"
        cf = "created_at >= NOW() - INTERVAL '30 days'"
    else:
        sf = "EXTRACT(YEAR FROM created_at)=EXTRACT(YEAR FROM NOW())"
        cf = "created_at >= NOW() - INTERVAL '365 days'"
    branch_filter = " AND branch_id=%s" if bid else ""
    bp = [bid] if bid else []
    try:
        with get_db() as conn:
            cur = conn.cursor()
            def q1(t,f,extra=""): cur.execute(f"SELECT COALESCE({t},0) as v FROM {f} WHERE user_id=%s AND {sf}{extra}", [oid]+bp); return cur.fetchone()["v"]
            revenue    = q1("SUM(total_amount)","sales",branch_filter)
            txn_count  = q1("COUNT(*)","sales",branch_filter)
            vat        = q1("SUM(vat_amount)","sales",branch_filter)
            disc       = q1("SUM(discount)","sales",branch_filter)
            expenses   = q1("SUM(amount)","expenses",branch_filter.replace("branch_id","branch_id") if bid else "")
            cur.execute("SELECT COUNT(*) as v FROM products WHERE user_id=%s"+(f" AND branch_id=%s" if bid else ""), [oid]+bp); total_products=cur.fetchone()["v"]
            cur.execute("SELECT COUNT(*) as v FROM products WHERE user_id=%s AND stock<=0"+(f" AND branch_id=%s" if bid else ""), [oid]+bp); out_stock=cur.fetchone()["v"]
            cur.execute("SELECT COUNT(*) as v FROM products WHERE user_id=%s AND stock>0 AND stock<=reorder"+(f" AND branch_id=%s" if bid else ""), [oid]+bp); low_stock=cur.fetchone()["v"]
            cur.execute("SELECT COUNT(*) as v FROM customers WHERE owner_id=%s", (oid,)); total_customers=cur.fetchone()["v"]
            cur.execute(f"SELECT DATE(created_at) as day,SUM(total_amount) as total FROM sales WHERE user_id=%s AND {cf}{'AND branch_id=%s' if bid else ''} GROUP BY DATE(created_at) ORDER BY day", [oid]+bp)
            chart_data = [{"day":str(r["day"]),"total":r["total"]} for r in cur.fetchall()]
            cur.execute("SELECT s.txn_id,s.customer,s.total_amount,s.created_at,si.name as item_name FROM sales s LEFT JOIN sale_items si ON si.sale_id=s.id WHERE s.user_id=%s ORDER BY s.created_at DESC LIMIT 5",(oid,))
            recent_sales=[dict(r) for r in cur.fetchall()]
            cur.execute("SELECT si.name,SUM(si.qty) as total_qty,SUM(si.subtotal) as total_rev FROM sale_items si JOIN sales s ON s.id=si.sale_id WHERE s.user_id=%s GROUP BY si.name ORDER BY total_qty DESC LIMIT 5",(oid,))
            top_products=[dict(r) for r in cur.fetchall()]
        return jsonify({"ok":True,"data":{
            "revenue":revenue,"txn_count":txn_count,"expenses":expenses,
            "net_profit":revenue-expenses,"vat_collected":vat,
            "discounts_given":disc,"total_products":total_products,
            "out_stock":out_stock,"low_stock":low_stock,
            "total_customers":total_customers,"chart_data":chart_data,
            "recent_sales":recent_sales,"top_products":top_products,
        }})
    except Exception as e:
        print(f"[DASHBOARD ERROR] {e}")
        return jsonify({"ok":False,"error":str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  EOD REPORT
# ══════════════════════════════════════════════════════════════════
@app.route("/api/eod-report", methods=["POST"])
@api_login_required
def eod_report():
    oid = get_owner_id()
    bid = get_active_branch()
    try:
        with get_db() as conn:
            cur = conn.cursor()
            bf  = " AND branch_id=%s" if bid else ""
            bp  = [bid] if bid else []
            cur.execute(f"SELECT COALESCE(SUM(total_amount),0) as v,COUNT(*) as c FROM sales WHERE user_id=%s AND DATE(created_at)=CURRENT_DATE{bf}", [oid]+bp)
            sr  = cur.fetchone()
            cur.execute(f"SELECT COALESCE(SUM(amount),0) as v FROM expenses WHERE user_id=%s AND DATE(created_at)=CURRENT_DATE{bf}", [oid]+bp)
            exp = cur.fetchone()["v"]
            cur.execute(f"SELECT COALESCE(SUM(vat_amount),0) as v FROM sales WHERE user_id=%s AND DATE(created_at)=CURRENT_DATE{bf}", [oid]+bp)
            vat = cur.fetchone()["v"]
            cur.execute(f"SELECT COALESCE(SUM(discount),0) as v FROM sales WHERE user_id=%s AND DATE(created_at)=CURRENT_DATE{bf}", [oid]+bp)
            disc= cur.fetchone()["v"]
            cur.execute(f"SELECT * FROM sales WHERE user_id=%s AND DATE(created_at)=CURRENT_DATE{bf} ORDER BY created_at DESC", [oid]+bp)
            recent=[dict(r) for r in cur.fetchall()]
            total_sales=sr["v"]; total_txns=sr["c"]
            cur.execute("""INSERT INTO eod_reports
                (owner_id,branch_id,report_date,total_sales,total_txns,total_expenses,vat_collected,discounts_given,net_profit)
                VALUES (%s,%s,CURRENT_DATE,%s,%s,%s,%s,%s,%s)""",
                (oid,bid,total_sales,total_txns,exp,vat,disc,total_sales-exp))
        return jsonify({"ok":True,"report":{
            "total_sales":total_sales,"total_txns":total_txns,
            "total_expenses":exp,"net_profit":total_sales-exp,
            "vat_collected":vat,"discounts_given":disc,"recent_sales":recent,
        }})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  REPORTS
# ══════════════════════════════════════════════════════════════════
@app.route("/api/reports")
@api_login_required
def get_report():
    oid   = get_owner_id()
    rtype = request.args.get("type","daily")
    year  = request.args.get("year",  date.today().year)
    month = request.args.get("month", date.today().month)
    rdate = request.args.get("date",  date.today().isoformat())
    try:
        with get_db() as conn:
            cur = conn.cursor()
            if rtype == "inventory":
                cur.execute("SELECT * FROM products WHERE user_id=%s ORDER BY category,name", (oid,))
                prods=cur.fetchall()
                total=len(prods); out=sum(1 for p in prods if p["stock"]<=0)
                low=sum(1 for p in prods if p["stock"]>0 and p["stock"]<=p["reorder"])
                data={"title":"Inventory Report","products":[dict(p) for p in prods],
                      "total_products":total,"in_stock":total-out,"low_stock":low,"out_stock":out}
            else:
                if rtype=="daily":   filt="DATE(s.created_at)=%s"; p=[rdate]
                elif rtype=="monthly": filt="EXTRACT(YEAR FROM s.created_at)=%s AND EXTRACT(MONTH FROM s.created_at)=%s"; p=[year,month]
                else: filt="s.created_at >= NOW() - INTERVAL '7 days'"; p=[]
                cur.execute(f"SELECT COALESCE(SUM(total_amount),0) as v FROM sales s WHERE user_id=%s AND {filt}",[oid]+p); rev=cur.fetchone()["v"]
                cur.execute(f"SELECT COUNT(*) as v FROM sales s WHERE user_id=%s AND {filt}",[oid]+p); txns=cur.fetchone()["v"]
                cur.execute(f"SELECT COALESCE(SUM(vat_amount),0) as v FROM sales s WHERE user_id=%s AND {filt}",[oid]+p); vat=cur.fetchone()["v"]
                cur.execute(f"SELECT COALESCE(SUM(discount),0) as v FROM sales s WHERE user_id=%s AND {filt}",[oid]+p); disc=cur.fetchone()["v"]
                cur.execute(f"SELECT COALESCE(SUM(amount),0) as v FROM expenses WHERE user_id=%s AND {filt.replace('s.created_at','created_at')}",[oid]+p); exp=cur.fetchone()["v"]
                cur.execute(f"SELECT si.name,SUM(si.qty) as total_qty,SUM(si.subtotal) as total_rev FROM sale_items si JOIN sales s ON s.id=si.sale_id WHERE s.user_id=%s AND {filt} GROUP BY si.name ORDER BY total_qty DESC LIMIT 10",[oid]+p)
                top=[dict(r) for r in cur.fetchall()]
                cur.execute(f"SELECT * FROM sales s WHERE s.user_id=%s AND {filt} ORDER BY s.created_at DESC LIMIT 20",[oid]+p)
                sales=[dict(r) for r in cur.fetchall()]
                labels={"daily":f"Daily Report — {rdate}","monthly":f"Monthly — {month}/{year}","weekly":"Weekly Report"}
                data={"title":labels.get(rtype,"Report"),"revenue":rev,"txn_count":txns,
                      "expenses":exp,"net_profit":rev-exp,"vat_collected":vat,
                      "discounts_given":disc,"top_products":top,"sales":sales}
        with get_db() as conn:
            cur=conn.cursor()
            cur.execute("INSERT INTO report_history (owner_id,report_type,delivered_via) VALUES (%s,%s,%s)",(oid,rtype,"downloaded"))
        return jsonify({"ok":True,"report":data})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500

@app.route("/api/reports/email", methods=["POST"])
@api_login_required
def email_report():
    oid    = get_owner_id()
    user   = get_current_user()
    d      = request.json or {}
    report = d.get("report",{})
    rtype  = d.get("type","daily")
    title  = report.get("title","Report")
    body   = f"<h2>{title}</h2><p>Revenue: KES {int(report.get('revenue',0)):,}</p><p>Net Profit: KES {int(report.get('net_profit',0)):,}</p>"
    ok     = send_email(user["email"], f"📊 {user['business_name']} — {title}", body)
    if ok:
        with get_db() as conn:
            cur=conn.cursor()
            cur.execute("INSERT INTO report_history (owner_id,report_type,delivered_via) VALUES (%s,%s,%s)",(oid,rtype,f"email:{user['email']}"))
    return jsonify({"ok":ok,"error":None if ok else "Email sending failed."})

@app.route("/api/reports/history")
@api_login_required
def report_history():
    oid = get_owner_id()
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute("SELECT * FROM report_history WHERE owner_id=%s ORDER BY created_at DESC LIMIT 20",(oid,))
        rows=cur.fetchall()
    return jsonify({"ok":True,"history":[dict(r) for r in rows]})


# ══════════════════════════════════════════════════════════════════
#  DATA EXPORT
# ══════════════════════════════════════════════════════════════════
@app.route("/api/export/<data_type>")
@api_login_required
def export_data(data_type):
    oid  = get_owner_id()
    fmt  = request.args.get("format","excel")
    with get_db() as conn:
        cur = conn.cursor()
        if data_type == "sales":
            cur.execute("""SELECT s.txn_id,s.created_at,s.customer,s.total_amount,
                s.vat_amount,s.discount,s.mpesa_code,s.price_tier,b.name as branch
                FROM sales s LEFT JOIN branches b ON b.id=s.branch_id
                WHERE s.user_id=%s ORDER BY s.created_at DESC""", (oid,))
            headers=["TXN ID","Date","Customer","Total (KES)","VAT (KES)","Discount (KES)","M-Pesa Code","Price Tier","Branch"]
            rows=[(r["txn_id"],str(r["created_at"]),r["customer"],r["total_amount"],r["vat_amount"],r["discount"],r["mpesa_code"],r["price_tier"],r["branch"] or "Main") for r in cur.fetchall()]
            fname="sales_export"
        elif data_type == "expenses":
            cur.execute("""SELECT e.created_at,e.description,e.category,e.amount,e.note,b.name as branch
                FROM expenses e LEFT JOIN branches b ON b.id=e.branch_id
                WHERE e.user_id=%s ORDER BY e.created_at DESC""", (oid,))
            headers=["Date","Description","Category","Amount (KES)","Note","Branch"]
            rows=[(str(r["created_at"]),r["description"],r["category"],r["amount"],r["note"],r["branch"] or "Main") for r in cur.fetchall()]
            fname="expenses_export"
        elif data_type == "inventory":
            cur.execute("SELECT name,category,barcode,price,wholesale_price,bulk_price,stock,reorder,discount_pct,discount_fixed FROM products WHERE user_id=%s ORDER BY category,name", (oid,))
            headers=["Product","Category","Barcode","Retail Price","Wholesale","Bulk Price","Stock","Reorder Level","Discount %","Fixed Discount"]
            rows=[(r["name"],r["category"],r["barcode"],r["price"],r["wholesale_price"],r["bulk_price"],r["stock"],r["reorder"],r["discount_pct"],r["discount_fixed"]) for r in cur.fetchall()]
            fname="inventory_export"
        elif data_type == "customers":
            cur.execute("SELECT name,phone,email,customer_type,loyalty_points,total_spent,created_at FROM customers WHERE owner_id=%s ORDER BY name", (oid,))
            headers=["Name","Phone","Email","Type","Loyalty Points","Total Spent (KES)","Joined"]
            rows=[(r["name"],r["phone"],r["email"],r["customer_type"],r["loyalty_points"],r["total_spent"],str(r["created_at"])) for r in cur.fetchall()]
            fname="customers_export"
        else:
            return jsonify({"ok":False,"error":"Unknown export type"}), 400

    if fmt == "csv":
        import csv, io as _io
        buf = _io.StringIO()
        w   = csv.writer(buf)
        w.writerow(headers)
        w.writerows(rows)
        return Response(buf.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition":f"attachment; filename={fname}.csv"})
    else:
        data = build_excel(headers, rows, data_type.title())
        if not data:
            return jsonify({"ok":False,"error":"openpyxl not installed"}), 500
        return Response(data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition":f"attachment; filename={fname}.xlsx"})


# ══════════════════════════════════════════════════════════════════
#  TEAM
# ══════════════════════════════════════════════════════════════════
@app.route("/api/team")
@owner_required
def get_team():
    user = get_current_user()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT u.id,u.name,u.email,u.phone,u.role,u.branch_id,u.created_at,
            b.name as branch_name FROM users u
            LEFT JOIN branches b ON b.id=u.branch_id
            WHERE u.owner_id=%s AND u.role='cashier' ORDER BY u.name""", (user["id"],))
        rows = cur.fetchall()
    return jsonify({"ok":True,"cashiers":[dict(r) for r in rows]})

@app.route("/api/team", methods=["POST"])
@owner_required
def add_cashier():
    user  = get_current_user()
    d     = request.json or {}
    email = (d.get("email")    or "").strip().lower()
    pwd   = (d.get("password") or "").strip()
    name  = (d.get("name")     or "").strip()
    bid   = d.get("branch_id")
    if not all([email,pwd,name]):
        return jsonify({"ok":False,"error":"Name, email and password required"}), 400
    if len(pwd) < 6:
        return jsonify({"ok":False,"error":"Password must be at least 6 characters"}), 400
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("""INSERT INTO users
                (email,password_hash,name,business_name,phone,role,owner_id,branch_id)
                VALUES (%s,%s,%s,%s,%s,'cashier',%s,%s) RETURNING id""",
                (email,hash_password(pwd),name,user["business_name"],
                 d.get("phone",""),user["id"],bid))
        return jsonify({"ok":True})
    except psycopg2.errors.UniqueViolation:
        return jsonify({"ok":False,"error":"Email already exists"}), 400

@app.route("/api/team/<int:cid>/reset-password", methods=["POST"])
@owner_required
def reset_cashier_password(cid):
    user = get_current_user()
    pwd  = (request.json or {}).get("new_password","")
    if len(pwd) < 6: return jsonify({"ok":False,"error":"Password too short"}), 400
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE users SET password_hash=%s WHERE id=%s AND owner_id=%s AND role='cashier'",
                    (hash_password(pwd),cid,user["id"]))
    return jsonify({"ok":True})

@app.route("/api/team/<int:cid>", methods=["DELETE"])
@owner_required
def remove_cashier(cid):
    user = get_current_user()
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM users WHERE id=%s AND owner_id=%s AND role='cashier'",
                    (cid,user["id"]))
    return jsonify({"ok":True})


# ══════════════════════════════════════════════════════════════════
#  HEALTH
# ══════════════════════════════════════════════════════════════════
@app.route("/health")
def health():
    try:
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) as c FROM users WHERE role='owner'")
            users = cur.fetchone()["c"]
        db_ok = True
    except Exception as e:
        db_ok = False; users = str(e)
    return jsonify({"ok":db_ok,"mpesa_env":MPESA_ENV,"callback":CALLBACK_URL,
                    "at_configured":bool(AT_API_KEY),
                    "email_configured":bool(GMAIL_USER),"users":users})


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
